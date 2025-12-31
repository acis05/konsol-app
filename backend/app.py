from __future__ import annotations

import io
import os
import re
import uuid
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Tuple

import pdfplumber
import openpyxl
from fastapi import Depends, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from sqlalchemy import Column, DateTime, String, create_engine, func
from sqlalchemy.orm import Session, declarative_base, sessionmaker
from sqlalchemy.types import JSON
from starlette.responses import JSONResponse, StreamingResponse

# Excel export
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# PDF export (ReportLab)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


# =========================
# App & CORS
# =========================
app = FastAPI(title="Konsolin • Mini Konsolidasi LK (PDF/Excel) + Arsip")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================
# Database (Postgres Railway)
# =========================
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
if not DATABASE_URL:
    DATABASE_URL = "sqlite:///./local.db"

if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)
Base = declarative_base()


class Report(Base):
    __tablename__ = "reports"
    report_id = Column(String(64), primary_key=True, index=True)
    created_at = Column(DateTime(timezone=True), server_default=func.now(), nullable=False)

    period_label = Column(String(128), nullable=True)
    as_of_date = Column(String(32), nullable=True)
    companies_text = Column(String, nullable=True)
    mapping_count = Column(String(16), nullable=True)

    payload = Column(JSON, nullable=False)
    result = Column(JSON, nullable=False)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@app.on_event("startup")
def on_startup():
    Base.metadata.create_all(bind=engine)


# =========================
# Helpers: Period detection
# =========================
ID_MONTHS = {
    "januari": 1, "jan": 1,
    "februari": 2, "feb": 2,
    "maret": 3, "mar": 3,
    "april": 4, "apr": 4,
    "mei": 5,
    "juni": 6, "jun": 6,
    "juli": 7, "jul": 7,
    "agustus": 8, "agu": 8, "ags": 8,
    "september": 9, "sep": 9,
    "oktober": 10, "okt": 10,
    "november": 11, "nov": 11,
    "desember": 12, "des": 12,
}


def detect_period_from_textlines(lines: List[str]) -> Dict[str, Optional[str]]:
    head = " \n ".join(lines[:80]).lower()

    m1 = re.search(r"\bper(?:\s+tgl\.?)?\s+(\d{1,2})\s+([a-zA-Z]{3,9})\s+(\d{4})\b", head)
    if m1:
        d = int(m1.group(1))
        mon = ID_MONTHS.get(m1.group(2).lower())
        y = int(m1.group(3))
        if mon:
            as_of = date(y, mon, d).isoformat()
            label = f"Per {d} {m1.group(2)} {y}"
            return {"label": label, "as_of": as_of}

    m2 = re.search(r"\bper\s+(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})\b", head)
    if m2:
        d, mon, y = int(m2.group(1)), int(m2.group(2)), int(m2.group(3))
        try:
            as_of = date(y, mon, d).isoformat()
            label = f"Per {d:02d}-{mon:02d}-{y}"
            return {"label": label, "as_of": as_of}
        except Exception:
            pass

    m3 = re.search(
        r"(dari\s+)?(\d{1,2})(?:[\/\-]|\s+)([a-zA-Z]{3,9}|\d{1,2})(?:[\/\-]|\s+)(\d{4}).{0,40}(s\.?d\.?|sampai|to|-\s*)\s*(\d{1,2})(?:[\/\-]|\s+)([a-zA-Z]{3,9}|\d{1,2})(?:[\/\-]|\s+)(\d{4})",
        head
    )
    if m3:
        d2 = int(m3.group(6))
        mon2_raw = m3.group(7)
        y2 = int(m3.group(8))
        try:
            if mon2_raw.isdigit():
                mon2 = int(mon2_raw)
            else:
                mon2 = ID_MONTHS.get(mon2_raw.lower())
            if mon2:
                as_of = date(y2, mon2, d2).isoformat()
                label = f"Periode s/d {d2:02d}-{mon2:02d}-{y2}"
                return {"label": label, "as_of": as_of}
        except Exception:
            pass

    return {"label": None, "as_of": None}


# =========================
# PDF parsing
# =========================
ACCOUNT_CODE_RE = re.compile(r"^\s*(\d{3,6}(?:\.\d+)?(?:-\d+)?[A-Za-z]?)\s+")
AMOUNT_RE = re.compile(r"(-?\(?\s*\d{1,3}(?:[.,]\d{3})*(?:[.,]\d+)?\s*\)?)\s*$")


def parse_amount_id(amount_str: str) -> Optional[int]:
    if amount_str is None:
        return None
    s = str(amount_str).strip()
    if not s:
        return None

    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1].strip()

    s = s.replace(" ", "")

    if "." in s and "," in s:
        s = s.replace(".", "")
        s = s.split(",")[0]
    else:
        s = s.replace(",", "")
        s = s.replace(".", "")

    if not s or not s.lstrip("-").isdigit():
        return None

    val = int(s)
    if neg:
        val = -abs(val)
    return val


def extract_lines_from_pdf(pdf_bytes: bytes) -> List[str]:
    lines: List[str] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for ln in text.splitlines():
                ln = ln.rstrip()
                if ln:
                    lines.append(ln)
    return lines


def parse_statement_rows_from_pdf(lines: List[str]) -> Tuple[List[Dict[str, Any]], List[str]]:
    rows: List[Dict[str, Any]] = []
    warnings: List[str] = []

    current_section: Optional[str] = None
    group_stack: List[str] = []

    SECTION_SET = {
        "ASET", "AKTIVA", "ASSET",
        "LIABILITAS", "KEWAJIBAN",
        "EKUITAS",
        "PENDAPATAN", "BEBAN", "LABA", "RUGI",
    }

    def is_heading(t: str) -> bool:
        if not t:
            return False
        if ACCOUNT_CODE_RE.match(t):
            return False
        if AMOUNT_RE.search(t):
            return False
        letters = sum(ch.isalpha() for ch in t)
        if letters < 3:
            return False
        up = t.upper()
        mostly_upper = (sum(ch.isupper() for ch in t if ch.isalpha()) / max(1, letters)) > 0.7
        has_keyword = any(k in up for k in ["ASET", "AKTIVA", "LIABILITAS", "EKUITAS", "PENDAPATAN", "BEBAN", "LANCAR"])
        return len(t) <= 70 and (mostly_upper or has_keyword)

    for ln in lines:
        t = ln.strip()
        up = t.upper()

        if up in SECTION_SET:
            current_section = t
            group_stack = [t]
            continue

        if is_heading(t):
            if current_section:
                if len(group_stack) >= 3:
                    group_stack = group_stack[:2] + [t]
                else:
                    if any(k in up for k in ["LANCAR", "TIDAK LANCAR", "JANGKA", "ASET", "LIABILITAS", "EKUITAS"]):
                        group_stack = [current_section, t] if current_section else [t]
                    else:
                        group_stack = group_stack + [t] if group_stack else [t]
            else:
                group_stack = [t]
            continue

        m_code = ACCOUNT_CODE_RE.match(t)
        if not m_code:
            continue
        m_amt = AMOUNT_RE.search(t)
        if not m_amt:
            continue

        code = m_code.group(1).strip()
        amt_raw = m_amt.group(1).strip()
        amt = parse_amount_id(amt_raw)
        if amt is None:
            warnings.append(f"Gagal parse amount: '{amt_raw}' pada line: {t}")
            continue

        body = t[m_code.end():].strip()
        if amt_raw in body:
            body = body[: body.rfind(amt_raw)].strip()

        gp = list(group_stack or ([] if not current_section else [current_section]))
        rows.append({
            "account_code": code,
            "account_name": body,
            "amount": int(amt),
            "section": current_section,
            "raw_line": t,
            "group_path": gp,
        })

    if not rows:
        warnings.append("PDF tidak terbaca dengan pola kode+amount. Jika punya Excel export Accurate, upload .xlsx (lebih akurat).")
    return rows, warnings


# =========================
# Excel parsing (ACCURATE exports)
# =========================
def _xlsx_sheet_to_matrix(xlsx_bytes: bytes) -> List[List[Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    matrix: List[List[Any]] = []
    for r in range(1, ws.max_row + 1):
        row = []
        for c in range(1, ws.max_column + 1):
            row.append(ws.cell(r, c).value)
        matrix.append(row)
    return matrix


def _find_header_row(matrix: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    def norm(x: Any) -> str:
        return (str(x).strip().lower() if x is not None else "")

    best = None

    for i, row in enumerate(matrix[:80]):
        norms = [norm(x) for x in row]
        if "kode akun" in norms:
            code_col = norms.index("kode akun")
            desc_col = None
            for j, v in enumerate(norms):
                if v == "deskripsi" or v == "keterangan":
                    desc_col = j
                    break
            if desc_col is None:
                continue

            parent_col = None
            for j, v in enumerate(norms):
                if v == "induk akun":
                    parent_col = j
                    break

            amount_col = None
            for j in range(desc_col + 1, min(desc_col + 6, len(norms))):
                if norms[j]:
                    if "nilai" in norms[j] or re.search(r"\d{1,2}\s*-\s*\d{1,2}", norms[j]) or re.search(r"\d{4}", norms[j]):
                        amount_col = j
                        break
            if amount_col is None:
                for j, v in enumerate(norms):
                    if "nilai" in v:
                        amount_col = j
                        break
            if amount_col is None:
                continue

            score = 10
            if parent_col is not None:
                score += 2

            if best is None or score > best[0]:
                best = (score, i, {"code": code_col, "desc": desc_col, "amount": amount_col, "parent": parent_col})

    if not best:
        raise ValueError("Header Excel tidak ditemukan. Pastikan file adalah export Accurate (ada kolom 'Kode Akun' & 'Deskripsi').")

    return best[1], best[2]


def parse_statement_rows_from_xlsx(xlsx_bytes: bytes) -> Tuple[List[Dict[str, Any]], List[str], Dict[str, Optional[str]]]:
    warnings: List[str] = []
    matrix = _xlsx_sheet_to_matrix(xlsx_bytes)

    top_lines: List[str] = []
    for r in matrix[:40]:
        for v in r[:6]:
            if isinstance(v, str) and v.strip():
                top_lines.append(v.strip())
    detected = detect_period_from_textlines(top_lines)

    header_i, colmap = _find_header_row(matrix)

    code_c = colmap["code"]
    desc_c = colmap["desc"]
    amt_c = colmap["amount"]
    parent_c = colmap.get("parent")

    current_section: Optional[str] = None
    group_stack: List[str] = []

    SECTION_HINTS = ("ASET", "AKTIVA", "ASSET", "LIABILITAS", "KEWAJIBAN", "EKUITAS", "PENDAPATAN", "BEBAN", "LABA", "RUGI")

    def is_footer(desc: str) -> bool:
        up = desc.upper()
        return ("ACCURATE" in up) or ("TERCETAK" in up) or ("HALAMAN" in up)

    def clean_str(x: Any) -> str:
        return str(x).strip() if x is not None else ""

    rows: List[Dict[str, Any]] = []

    for r in matrix[header_i + 1:]:
        code = r[code_c] if code_c < len(r) else None
        desc = clean_str(r[desc_c] if desc_c < len(r) else None)
        amt = r[amt_c] if amt_c < len(r) else None
        parent = clean_str(r[parent_c] if (parent_c is not None and parent_c < len(r)) else None)

        if desc and is_footer(desc):
            break

        code_str = ""
        if code is not None:
            code_str = str(code).strip()
            if re.fullmatch(r"\d+(\.0+)?", code_str):
                code_str = str(int(float(code_str)))

        amt_val = None
        if amt is not None and str(amt).strip() != "":
            try:
                amt_val = int(round(float(amt)))
            except Exception:
                amt_val = parse_amount_id(str(amt))

        if (not code_str) and desc and (amt_val is None):
            up = desc.upper()
            if any(k in up for k in SECTION_HINTS):
                current_section = desc
                group_stack = [desc]
            else:
                if current_section:
                    if len(group_stack) >= 3:
                        group_stack = group_stack[:2] + [desc]
                    else:
                        group_stack.append(desc)
                else:
                    group_stack = [desc]
            continue

        if code_str and amt_val is not None:
            gp = list(group_stack)
            if not gp and current_section:
                gp = [current_section]

            rows.append({
                "account_code": code_str,
                "account_name": desc,
                "amount": int(amt_val),
                "section": current_section,
                "raw_line": f"{code_str} {desc} {amt_val}",
                "group_path": gp,
                "is_parent_account": (parent.lower() == "ya"),
            })
            continue

        continue

    if not rows:
        warnings.append("Tidak ada baris akun dari Excel terdeteksi. Cek apakah sheet pertama berisi tabel export Accurate.")
    return rows, warnings, detected


# =========================
# Models (API)
# =========================
class ParsedRow(BaseModel):
    account_code: str
    account_name: str
    amount: int
    section: Optional[str] = None
    raw_line: Optional[str] = None
    group_path: List[str] = Field(default_factory=list)


class CompanyPayload(BaseModel):
    company_name: str
    period: Optional[str] = None
    bs_rows: List[ParsedRow] = Field(default_factory=list)
    is_rows: List[ParsedRow] = Field(default_factory=list)


class PairMapping(BaseModel):
    pair_name: str
    company_ar: str
    ar_account_code: str
    company_ap: str
    ap_account_code: str
    note: Optional[str] = None


class ConsolidateOptions(BaseModel):
    elim_method: str = "MIN_ABS"
    strict_match: bool = False

    include_details: bool = True


class ConsolidateRequest(BaseModel):
    companies: List[CompanyPayload]
    pair_mappings: List[PairMapping] = Field(default_factory=list)
    options: ConsolidateOptions = Field(default_factory=ConsolidateOptions)


# =========================
# Consolidation logic
# =========================
def index_balances(rows: List[ParsedRow]) -> Dict[str, ParsedRow]:
    return {r.account_code: r for r in rows}


def union_accounts_meta(companies: List[CompanyPayload], statement: str) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for c in companies:
        rlist = c.bs_rows if statement == "BS" else c.is_rows
        for r in rlist:
            if r.account_code not in out:
                out[r.account_code] = {"name": r.account_name, "group_path": list(r.group_path or [])}
            else:
                if (not out[r.account_code].get("group_path")) and (r.group_path):
                    out[r.account_code]["group_path"] = list(r.group_path or [])
    return out


def build_company_amount_map(companies: List[CompanyPayload], statement: str) -> Dict[str, Dict[str, int]]:
    res: Dict[str, Dict[str, int]] = {}
    for c in companies:
        rlist = c.bs_rows if statement == "BS" else c.is_rows
        for r in rlist:
            res.setdefault(r.account_code, {})
            res[r.account_code][c.company_name] = res[r.account_code].get(c.company_name, 0) + int(r.amount)
    return res


def _norm_label(s: str) -> str:
    s = (s or "").strip().upper()
    s = re.sub(r"\s+", " ", s)
    return s


def _sum_dicts(a: Dict[str, int], b: Dict[str, int]) -> Dict[str, int]:
    out = dict(a or {})
    for k, v in (b or {}).items():
        out[k] = out.get(k, 0) + int(v or 0)
    return out


def consolidate(companies: List[CompanyPayload], mappings: List[PairMapping], options: ConsolidateOptions):
    bs_meta = union_accounts_meta(companies, "BS")
    is_meta = union_accounts_meta(companies, "IS")
    bs_by_company = build_company_amount_map(companies, "BS")
    is_by_company = build_company_amount_map(companies, "IS")

    elim_effect_bs: Dict[str, int] = {}
    elimination_journal: List[Dict[str, Any]] = []
    unreconciled: List[Dict[str, Any]] = []

    company_map = {c.company_name: c for c in companies}

    for mp in mappings:
        c_ar = company_map.get(mp.company_ar)
        c_ap = company_map.get(mp.company_ap)
        if not c_ar or not c_ap:
            unreconciled.append({"pair_name": mp.pair_name, "error": "Company not found", **mp.model_dump()})
            continue

        ar_idx = index_balances(c_ar.bs_rows)
        ap_idx = index_balances(c_ap.bs_rows)
        ar_row = ar_idx.get(mp.ar_account_code)
        ap_row = ap_idx.get(mp.ap_account_code)

        if not ar_row or not ap_row:
            unreconciled.append({
                "pair_name": mp.pair_name,
                "company_ar": mp.company_ar,
                "ar_account_code": mp.ar_account_code,
                "company_ap": mp.company_ap,
                "ap_account_code": mp.ap_account_code,
                "error": "AR/AP code not found in BS rows",
            })
            continue

        ar_bal = int(ar_row.amount)
        ap_bal = int(ap_row.amount)

        elim_amt = min(abs(ar_bal), abs(ap_bal))
        diff = abs(ar_bal) - abs(ap_bal)

        elim_effect_bs[mp.ar_account_code] = elim_effect_bs.get(mp.ar_account_code, 0) - elim_amt
        elim_effect_bs[mp.ap_account_code] = elim_effect_bs.get(mp.ap_account_code, 0) - elim_amt

        elimination_journal.append({
            "pair_name": mp.pair_name,
            "amount": elim_amt,
            "ar_balance": ar_bal,
            "ap_balance": ap_bal,
            "difference": diff,
            "status": "MATCH" if diff == 0 else "MISMATCH",
            "lines": [
                {"company": mp.company_ap, "drcr": "DR", "account_code": mp.ap_account_code, "account_name": ap_row.account_name, "amount": elim_amt},
                {"company": mp.company_ar, "drcr": "CR", "account_code": mp.ar_account_code, "account_name": ar_row.account_name, "amount": elim_amt},
            ],
        })

        if diff != 0:
            unreconciled.append({
                "pair_name": mp.pair_name,
                "company_ar": mp.company_ar,
                "ar_account_code": mp.ar_account_code,
                "company_ap": mp.company_ap,
                "ap_account_code": mp.ap_account_code,
                "ar_balance": ar_bal,
                "ap_balance": ap_bal,
                "difference": diff,
            })

    def build_hier(meta: Dict[str, Dict[str, Any]], by_company: Dict[str, Dict[str, int]], elim_effect: Optional[Dict[str, int]] = None, statement: str = "BS"):
        # 1) accounts list
        accounts: List[Dict[str, Any]] = []
        for code, m in sorted(meta.items(), key=lambda x: x[0]):
            name = (m.get("name") or "")
            gp = list(m.get("group_path") or [])
            bc = by_company.get(code, {}) or {}
            total_before = sum(int(v or 0) for v in bc.values())
            elimination = int((elim_effect or {}).get(code, 0))
            total_after = total_before + elimination
            accounts.append({
                "type": "ACCOUNT",
                "group_path": gp,
                "level": max(1, len(gp) + 1),
                "account_code": code,
                "account_name": name,
                "by_company": bc,
                "total_before": total_before,
                "elimination": elimination,
                "total_after": total_after,
            })

        # 2) group_totals by prefix
        group_totals: Dict[Tuple[str, ...], Dict[str, Any]] = {}

        def add_to_group(key: Tuple[str, ...], label: str, acc: Dict[str, Any]):
            g = group_totals.get(key)
            if not g:
                g = {
                    "type": "GROUP",
                    "key": key,
                    "level": len(key),
                    "label": label,
                    "by_company": {},
                    "elimination": 0,
                    "total_after": 0,
                }
                group_totals[key] = g

            for cn, v in (acc.get("by_company") or {}).items():
                g["by_company"][cn] = g["by_company"].get(cn, 0) + int(v or 0)

            g["elimination"] += int(acc.get("elimination", 0) or 0)
            g["total_after"] += int(acc.get("total_after", 0) or 0)

        for acc in accounts:
            gp = acc.get("group_path") or []
            for i in range(1, len(gp) + 1):
                key = tuple(gp[:i])
                add_to_group(key, gp[i - 1], acc)

        # 3) total labels mapping (based on group label)
        # NOTE: we match by normalized label
        if statement == "BS":
            want_group_totals = {
                "ASET LANCAR": "Total Aktiva Lancar",
                "AKTIVA LANCAR": "Total Aktiva Lancar",
                "ASSET LANCAR": "Total Aktiva Lancar",

                "ASET TETAP": "Total Aktiva Tetap",
                "AKTIVA TETAP": "Total Aktiva Tetap",
                "ASET TIDAK LANCAR": "Total Aktiva Tetap",
                "AKTIVA TIDAK LANCAR": "Total Aktiva Tetap",
                "ASET NON LANCAR": "Total Aktiva Tetap",

                "LIABILITAS JANGKA PENDEK": "Total Liabilitas Jangka Pendek",
                "KEWAJIBAN JANGKA PENDEK": "Total Liabilitas Jangka Pendek",

                "LIABILITAS JANGKA PANJANG": "Total Liabilitas Jangka Panjang",
                "KEWAJIBAN JANGKA PANJANG": "Total Liabilitas Jangka Panjang",

                "EKUITAS": "Total Ekuitas",
            }
        else:
            want_group_totals = {
                "PENDAPATAN": "Total Pendapatan",
                "PENJUALAN": "Total Pendapatan",

                "HARGA POKOK PENJUALAN": "Total Harga Pokok Penjualan",
                "HPP": "Total Harga Pokok Penjualan",

                "BEBAN OPERASIONAL": "Total Beban Operasional",
                "BEBAN USAHA": "Total Beban Operasional",

                "PENDAPATAN DI LUAR USAHA": "Total Pendapatan Diluar Usaha",
                "PENDAPATAN DILUAR USAHA": "Total Pendapatan Diluar Usaha",
                "PENDAPATAN LAIN-LAIN": "Total Pendapatan Diluar Usaha",
                "PENDAPATAN LAIN LAIN": "Total Pendapatan Diluar Usaha",

                "BEBAN DI LUAR USAHA": "Total Beban Diluar Usaha",
                "BEBAN DILUAR USAHA": "Total Beban Diluar Usaha",
                "BEBAN LAIN-LAIN": "Total Beban Diluar Usaha",
                "BEBAN LAIN LAIN": "Total Beban Diluar Usaha",
            }

        def make_total_row(label: str, level: int, g: Dict[str, Any]) -> Dict[str, Any]:
            return {
                "type": "TOTAL",
                "level": int(level),
                "label": label,
                "by_company": dict(g.get("by_company") or {}),
                "elimination": int(g.get("elimination") or 0),
                "total_after": int(g.get("total_after") or 0),
            }

        # 4) emit in order: open group rows, optional accounts, then close groups with TOTAL lines
        accounts_sorted = sorted(accounts, key=lambda a: ((a.get("group_path") or []), a.get("account_code") or ""))

        out: List[Dict[str, Any]] = []
        emitted_open: set[Tuple[str, ...]] = set()
        open_stack: List[Tuple[str, ...]] = []

        def close_groups_until(common_len: int):
            # close from deepest to common_len
            while len(open_stack) > common_len:
                key = open_stack.pop()
                g = group_totals.get(key)
                if not g:
                    continue
                lbl = _norm_label(g.get("label", ""))
                total_name = want_group_totals.get(lbl)
                if total_name:
                    out.append(make_total_row(total_name, len(key), g))

        for acc in accounts_sorted:
            gp: List[str] = acc.get("group_path") or []
            keys = [tuple(gp[:i]) for i in range(1, len(gp) + 1)]

            # compute common prefix length between current open_stack and new keys
            common = 0
            for i in range(min(len(open_stack), len(keys))):
                if open_stack[i] == keys[i]:
                    common += 1
                else:
                    break

            close_groups_until(common)

            # open any new groups
            for i in range(common, len(keys)):
                key = keys[i]
                if key not in emitted_open:
                    g = group_totals.get(key)
                    if g:
                        out.append({
                            "type": "GROUP",
                            "level": g["level"],
                            "label": g["label"],
                            "by_company": g["by_company"],
                            "elimination": g["elimination"],
                            "total_after": g["total_after"],
                        })
                    emitted_open.add(key)
                # ensure stack aligned
                if len(open_stack) <= i:
                    open_stack.append(key)

            if options.include_details:
                out.append(acc)

        close_groups_until(0)

        # 5) compute overall totals (BS + IS formulas)
        # helpers to grab totals by desired label name if we already emitted it
        def find_total_value(label_text: str) -> Optional[Dict[str, Any]]:
            for r in reversed(out):
                if r.get("type") == "TOTAL" and _norm_label(r.get("label", "")) == _norm_label(label_text):
                    return r
            return None

        # compute from group_totals (fallback)
        def group_total_by_label(label_candidates: List[str]) -> Optional[Dict[str, Any]]:
            cset = {_norm_label(x) for x in label_candidates}
            # choose the deepest matching group key to avoid overly broad
            best_key = None
            for key, g in group_totals.items():
                if _norm_label(g.get("label", "")) in cset:
                    if best_key is None or len(key) > len(best_key):
                        best_key = key
            if best_key is None:
                return None
            return group_totals.get(best_key)

        def make_formula_total(label: str, level: int, byc: Dict[str, int], elim: int, total_after: int) -> Dict[str, Any]:
            return {
                "type": "TOTAL",
                "level": int(level),
                "label": label,
                "by_company": dict(byc or {}),
                "elimination": int(elim or 0),
                "total_after": int(total_after or 0),
            }

        if statement == "BS":
            # Total Aktiva = total ASET/AKTIVA top-level if exists, else sum of all accounts
            g_aset = group_total_by_label(["ASET", "AKTIVA", "ASSET"])
            if g_aset:
                out.append(make_total_row("Total Aktiva", 1, g_aset))

            g_liab = group_total_by_label(["LIABILITAS", "KEWAJIBAN"])
            if g_liab:
                out.append(make_total_row("Total Liabilitas", 1, g_liab))

            g_ekuitas = group_total_by_label(["EKUITAS"])
            if g_ekuitas:
                out.append(make_total_row("Total Ekuitas", 1, g_ekuitas))

            # Total Passiva = Liabilitas + Ekuitas
            if g_liab and g_ekuitas:
                byc = _sum_dicts(g_liab.get("by_company") or {}, g_ekuitas.get("by_company") or {})
                elim = int(g_liab.get("elimination") or 0) + int(g_ekuitas.get("elimination") or 0)
                total_after = int(g_liab.get("total_after") or 0) + int(g_ekuitas.get("total_after") or 0)
                out.append(make_formula_total("Total Passiva", 1, byc, elim, total_after))

        else:
            # IS formulas
            t_pend = find_total_value("Total Pendapatan")
            if not t_pend:
                g = group_total_by_label(["PENDAPATAN", "PENJUALAN"])
                if g:
                    t_pend = make_total_row("Total Pendapatan", 1, g)
                    out.append(t_pend)

            t_hpp = find_total_value("Total Harga Pokok Penjualan")
            if not t_hpp:
                g = group_total_by_label(["HARGA POKOK PENJUALAN", "HPP"])
                if g:
                    t_hpp = make_total_row("Total Harga Pokok Penjualan", 1, g)
                    out.append(t_hpp)

            # Laba Kotor = Pendapatan - HPP
            if t_pend and t_hpp:
                byc = {}
                for k in set((t_pend.get("by_company") or {}).keys()) | set((t_hpp.get("by_company") or {}).keys()):
                    byc[k] = int((t_pend.get("by_company") or {}).get(k, 0)) - int((t_hpp.get("by_company") or {}).get(k, 0))
                elim = int(t_pend.get("elimination") or 0) - int(t_hpp.get("elimination") or 0)
                total_after = int(t_pend.get("total_after") or 0) - int(t_hpp.get("total_after") or 0)
                out.append(make_formula_total("Laba Kotor", 1, byc, elim, total_after))

            t_bop = find_total_value("Total Beban Operasional")
            if not t_bop:
                g = group_total_by_label(["BEBAN OPERASIONAL", "BEBAN USAHA"])
                if g:
                    t_bop = make_total_row("Total Beban Operasional", 1, g)
                    out.append(t_bop)

            # Laba Operasional = Laba Kotor - Beban Operasional
            t_lk = find_total_value("Laba Kotor")
            if t_lk and t_bop:
                byc = {}
                for k in set((t_lk.get("by_company") or {}).keys()) | set((t_bop.get("by_company") or {}).keys()):
                    byc[k] = int((t_lk.get("by_company") or {}).get(k, 0)) - int((t_bop.get("by_company") or {}).get(k, 0))
                elim = int(t_lk.get("elimination") or 0) - int(t_bop.get("elimination") or 0)
                total_after = int(t_lk.get("total_after") or 0) - int(t_bop.get("total_after") or 0)
                out.append(make_formula_total("Laba Operasional", 1, byc, elim, total_after))

            t_pdu = find_total_value("Total Pendapatan Diluar Usaha")
            if not t_pdu:
                g = group_total_by_label(["PENDAPATAN DI LUAR USAHA", "PENDAPATAN DILUAR USAHA", "PENDAPATAN LAIN-LAIN", "PENDAPATAN LAIN LAIN"])
                if g:
                    t_pdu = make_total_row("Total Pendapatan Diluar Usaha", 1, g)
                    out.append(t_pdu)

            t_bdu = find_total_value("Total Beban Diluar Usaha")
            if not t_bdu:
                g = group_total_by_label(["BEBAN DI LUAR USAHA", "BEBAN DILUAR USAHA", "BEBAN LAIN-LAIN", "BEBAN LAIN LAIN"])
                if g:
                    t_bdu = make_total_row("Total Beban Diluar Usaha", 1, g)
                    out.append(t_bdu)

            # Laba Bersih = Laba Operasional + Pendapatan Luar - Beban Luar
            t_lo = find_total_value("Laba Operasional")
            if t_lo:
                byc = dict(t_lo.get("by_company") or {})
                elim = int(t_lo.get("elimination") or 0)
                total_after = int(t_lo.get("total_after") or 0)

                if t_pdu:
                    byc = _sum_dicts(byc, t_pdu.get("by_company") or {})
                    elim += int(t_pdu.get("elimination") or 0)
                    total_after += int(t_pdu.get("total_after") or 0)

                if t_bdu:
                    for k, v in (t_bdu.get("by_company") or {}).items():
                        byc[k] = byc.get(k, 0) - int(v or 0)
                    elim -= int(t_bdu.get("elimination") or 0)
                    total_after -= int(t_bdu.get("total_after") or 0)

                out.append(make_formula_total("Laba Bersih", 1, byc, elim, total_after))

        return out

    return {
        "bs_comparison": build_hier(bs_meta, bs_by_company, elim_effect_bs, statement="BS"),
        "is_comparison": build_hier(is_meta, is_by_company, None, statement="IS"),
        "elimination_journal": elimination_journal,
        "unreconciled": unreconciled,
    }


# =========================
# Routes: Parse + Consolidate
# =========================
@app.post("/api/parse")
async def api_parse(
    company_name: str = Form(...),
    period: str = Form(""),
    statement: str = Form(...),  # BS / IS
    file: UploadFile = File(...)
):
    raw = await file.read()
    filename = (file.filename or "").lower()
    ctype = (file.content_type or "").lower()

    if filename.endswith(".xlsx") or "spreadsheet" in ctype or "excel" in ctype:
        try:
            rows, warnings, detected = parse_statement_rows_from_xlsx(raw)
            return JSONResponse({
                "company_name": company_name,
                "period": period or None,
                "statement": statement,
                "detected_period": detected,
                "rows": rows,
                "warnings": warnings,
                "source_type": "XLSX",
            })
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Gagal parse Excel: {e}")

    try:
        lines = extract_lines_from_pdf(raw)
        detected = detect_period_from_textlines(lines)
        rows, warnings = parse_statement_rows_from_pdf(lines)
        return JSONResponse({
            "company_name": company_name,
            "period": period or None,
            "statement": statement,
            "detected_period": detected,
            "rows": rows,
            "warnings": warnings,
            "source_type": "PDF",
        })
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal parse PDF: {e}")


@app.post("/api/consolidate")
async def api_consolidate(req: ConsolidateRequest):
    return JSONResponse(consolidate(req.companies, req.pair_mappings, req.options))


# =========================
# Routes: Reports (Arsip)
# =========================
class CreateReportRequest(BaseModel):
    companies: List[CompanyPayload]
    pair_mappings: List[PairMapping] = Field(default_factory=list)
    options: ConsolidateOptions = Field(default_factory=ConsolidateOptions)
    period_label: Optional[str] = None
    as_of: Optional[str] = None


@app.post("/api/reports")
def create_report(req: CreateReportRequest, db: Session = Depends(get_db)):
    result = consolidate(req.companies, req.pair_mappings, req.options)
    rid = "rpt_" + datetime.utcnow().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:6]
    companies_text = "|".join([c.company_name for c in req.companies])

    r = Report(
        report_id=rid,
        period_label=req.period_label,
        as_of_date=req.as_of,
        companies_text=companies_text,
        mapping_count=str(len(req.pair_mappings)),
        payload=req.model_dump(),
        result=result,
    )
    db.add(r)
    db.commit()

    return {"report_id": rid, "created_at": datetime.utcnow().isoformat() + "Z"}


@app.get("/api/reports")
def list_reports(query: Optional[str] = None, limit: int = 50, db: Session = Depends(get_db)):
    q = db.query(Report).order_by(Report.created_at.desc())
    if query:
        like = f"%{query.lower()}%"
        q = q.filter(
            func.lower(Report.companies_text).like(like)
            | func.lower(func.coalesce(Report.period_label, "")).like(like)
        )
    items = q.limit(min(max(limit, 1), 200)).all()

    return {
        "items": [
            {
                "report_id": r.report_id,
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "period_label": r.period_label,
                "as_of": r.as_of_date,
                "companies": (r.companies_text.split("|") if r.companies_text else []),
                "mapping_count": int(r.mapping_count or "0"),
            }
            for r in items
        ]
    }


@app.get("/api/reports/{report_id}")
def get_report(report_id: str, db: Session = Depends(get_db)):
    r = db.query(Report).filter(Report.report_id == report_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Report not found")
    return {
        "report_id": r.report_id,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "period_label": r.period_label,
        "as_of": r.as_of_date,
        "payload": r.payload,
        "result": r.result,
    }


@app.delete("/api/reports/{report_id}")
def delete_report(report_id: str, db: Session = Depends(get_db)):
    r = db.query(Report).filter(Report.report_id == report_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Report not found")
    db.delete(r)
    db.commit()
    return {"ok": True}


# =========================
# Export helpers
# =========================
def _company_names(req: ConsolidateRequest) -> List[str]:
    return [c.company_name for c in req.companies]


def _fmt_id(n: int) -> str:
    s = f"{int(n):,}"
    return s.replace(",", ".")


def _make_excel(req: ConsolidateRequest) -> bytes:
    result = consolidate(req.companies, req.pair_mappings, req.options)
    companies = _company_names(req)

    wb = Workbook()

    def add_sheet(title: str, rows: List[Dict[str, Any]]):
        ws = wb.create_sheet(title)

        headers = ["Level", "Tipe", "Kode", "Nama Akun"] + companies + ["Eliminasi", "Total Konsol"]
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

        for r in rows:
            rtype = r.get("type", "ACCOUNT")
            level = int(r.get("level", 1) or 1)
            code = r.get("account_code", "") if rtype == "ACCOUNT" else ""
            name = r.get("account_name", "") if rtype == "ACCOUNT" else (r.get("label", "") or "")
            name = ("  " * max(0, level - 1)) + name

            byc = r.get("by_company", {}) or {}
            line = [level, rtype, code, name]
            for cn in companies:
                line.append(int(byc.get(cn, 0) or 0))
            line.append(int(r.get("elimination", 0) or 0))
            line.append(int(r.get("total_after", 0) or 0))
            ws.append(line)

        for row in range(2, ws.max_row + 1):
            for col in range(5, len(headers) + 1):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")

        widths = [8, 10, 14, 50] + [18] * len(companies) + [16, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    default = wb.active
    wb.remove(default)

    add_sheet("Neraca_Konsol", result.get("bs_comparison", []))
    add_sheet("LabaRugi_Konsol", result.get("is_comparison", []))

    ws = wb.create_sheet("Jurnal_Eliminasi")
    ws.append(["Pair", "DR", "CR", "Elim", "Selisih", "Status"])
    for col in range(1, 7):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    for j in (result.get("elimination_journal") or []):
        dr = (j.get("lines") or [{}])[0].get("account_code", "")
        cr = (j.get("lines") or [{}, {}])[1].get("account_code", "")
        ws.append([
            j.get("pair_name", ""),
            dr,
            cr,
            int(j.get("amount", 0) or 0),
            int(j.get("difference", 0) or 0),
            j.get("status", "")
        ])

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _table_style_clear() -> TableStyle:
    return TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f2f4f7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#111827")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])


def _make_pdf(req: ConsolidateRequest) -> bytes:
    result = consolidate(req.companies, req.pair_mappings, req.options)
    companies = _company_names(req)

    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm
    )

    styles = getSampleStyleSheet()
    title_style = styles["Heading2"]
    title_style.fontName = "Helvetica-Bold"
    title_style.fontSize = 13

    normal = styles["Normal"]
    normal.fontName = "Helvetica"
    normal.fontSize = 8.5
    normal.leading = 10

    def build_table(title: str, rows: List[Dict[str, Any]]):
        story_part: List[Any] = [Paragraph(title, title_style), Spacer(1, 6)]

        headers = ["Kode", "Nama Akun"] + companies + ["Elim", "Total"]
        data: List[List[Any]] = [headers]

        for r in rows:
            rtype = r.get("type", "ACCOUNT")
            level = int(r.get("level", 1) or 1)

            if rtype in ("GROUP", "TOTAL"):
                code = ""
                name = r.get("label", "") or ""
                name = ("&nbsp;" * 2 * max(0, level - 1)) + f"<b>{name}</b>"
                name_cell = Paragraph(name, normal)
            else:
                code = r.get("account_code", "") or ""
                name = r.get("account_name", "") or ""
                name = ("&nbsp;" * 2 * max(0, level - 1)) + name
                name_cell = Paragraph(name, normal)

            byc = r.get("by_company", {}) or {}
            row_line: List[Any] = [code, name_cell]
            for cn in companies:
                row_line.append(_fmt_id(byc.get(cn, 0) or 0))
            row_line.append(_fmt_id(r.get("elimination", 0) or 0))
            row_line.append(_fmt_id(r.get("total_after", 0) or 0))
            data.append(row_line)

        col_widths = [28 * mm, 110 * mm] + [26 * mm] * len(companies) + [24 * mm, 28 * mm]
        tbl = Table(data, colWidths=col_widths, repeatRows=1)

        st = _table_style_clear()
        st.add("ALIGN", (0, 0), (1, -1), "LEFT")
        st.add("ALIGN", (2, 0), (-1, -1), "RIGHT")

        for i in range(1, len(data)):
            if rows[i - 1].get("type") == "GROUP":
                st.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#fff3ea"))
            if rows[i - 1].get("type") == "TOTAL":
                st.add("BACKGROUND", (0, i), (-1, i), colors.HexColor("#fff0d9"))
                st.add("LINEABOVE", (0, i), (-1, i), 1, colors.HexColor("#f59e0b"))
        tbl.setStyle(st)

        story_part.append(tbl)
        return story_part

    story: List[Any] = []
    story += build_table("Neraca Konsolidasi (Komparasi)", result.get("bs_comparison", []))
    story.append(PageBreak())
    story += build_table("Laba/Rugi Konsolidasi (Komparasi)", result.get("is_comparison", []))
    story.append(PageBreak())

    story.append(Paragraph("Jurnal Eliminasi", title_style))
    story.append(Spacer(1, 6))

    je = result.get("elimination_journal", []) or []
    je_headers = ["Pair", "DR", "CR", "Elim", "Selisih", "Status"]
    je_data: List[List[Any]] = [je_headers]

    for j in je:
        dr = (j.get("lines") or [{}])[0].get("account_code", "")
        cr = (j.get("lines") or [{}, {}])[1].get("account_code", "")
        je_data.append([
            j.get("pair_name", ""),
            dr,
            cr,
            _fmt_id(j.get("amount", 0) or 0),
            _fmt_id(j.get("difference", 0) or 0),
            j.get("status", "")
        ])

    je_tbl = Table(
        je_data,
        colWidths=[90 * mm, 28 * mm, 28 * mm, 25 * mm, 25 * mm, 22 * mm],
        repeatRows=1
    )

    je_style = _table_style_clear()
    je_style.add("ALIGN", (0, 0), (2, -1), "LEFT")
    je_style.add("ALIGN", (3, 0), (4, -1), "RIGHT")
    je_tbl.setStyle(je_style)

    story.append(je_tbl)

    doc.build(story)
    return bio.getvalue()


# =========================
# Export routes
# =========================
@app.post("/api/export/excel")
def export_excel(req: ConsolidateRequest):
    data = _make_excel(req)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=konsolidasi.xlsx"},
    )


@app.post("/api/export/pdf")
def export_pdf(req: ConsolidateRequest):
    data = _make_pdf(req)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=konsolidasi.pdf"},
    )